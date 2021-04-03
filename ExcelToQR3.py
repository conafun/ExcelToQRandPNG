#encoding:utf-8
from PIL import Image
import qrcode
import pygame
from openpyxl import load_workbook
pygame.init()  #初始化pygame

for i in range(1, 4):
    #i_str = str (i)
# 读取excel单元格内容
    wb=load_workbook('Sample.xlsx')
    sheet=wb.worksheets[0]
    sh=sheet.cell(i, 1).value #(x,y)代表x行，y列
    sh2=sheet.cell(i, 2).value
    sh3=sheet.cell(i, 3).value
    sh4=sheet.cell(i, 4).value
    sh5=sheet.cell(i, 5).value
    sh6=sheet.cell(i, 6).value
    sh7=sheet.cell(i, 7).value
    sh8=sheet.cell(i, 8).value
    tittle=sh+sh2
    detail=sh+'\n'+sh2+'\n'+sh3+'\n'+sh4+'\n'+sh5+'\n'+sh6+'\n'+sh7
#把excel内读取到的文字变成图片，生成二维码抬头
    font=pygame.font.SysFont('FangSong',26)
    ftext=font.render(tittle, True, (0, 0, 0),(255, 255, 255))
    pygame.image.save(ftext, sh+'.jpg')
#根据需要，收集不同单元格内信息，生产QR码
    icon = Image.open(sh+'.jpg')
    img=qrcode.make(detail)
    img2=img.paste(icon,(0, 0), mask=None)
#中心logo的获取
    icon1 = Image.open("logo.png")
    # 获取图片的宽高
    img_w, img_h = img.size
    factor = 3
    size_w = int(img_w / factor)
    size_h = int(img_h / factor)
    icon1_w, icon1_h = icon1.size
    if icon1_w > size_w:
        icon1_w = size_w
    if icon1_h > size_h:
        icon1_h = size_h
    # 重新设置logo的尺寸
    icon1 = icon1.resize((icon1_w, icon1_h), Image.ANTIALIAS)
    w = int((img_w - icon1_w) / 2)
    h = int((img_h - icon1_h) / 2)
    img.paste(icon1, (w, h), icon1)

    #with open(i_str+'.png', 'wb') as f:
    with open(sh+'.png', 'wb') as f:
        img.save(f)
    i=i+1