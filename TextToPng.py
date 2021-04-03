import pygame
pygame.init()  #初始化pygame
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

def detail(i):
  wb = load_workbook('Sample.xlsx')
  sheet = wb.worksheets[0]
  sh = sheet.cell(i, 1).value  # (x,y)代表x行，y列
  sh2 = sheet.cell(i, 2).value
  sh3 = sheet.cell(i, 3).value
  sh4 = sheet.cell(i, 4).value
  sh5 = sheet.cell(i, 5).value
  sh6 = sheet.cell(i, 6).value
  sh7 = sheet.cell(i, 7).value

  detail = sh + '\n' + sh2 + '\n' + sh3 + '\n' + sh4 + '\n' + sh5 + '\n' + sh6 + '\n' + sh7
  return detail

class ImgText:
  font = ImageFont.truetype("simsun.ttc", 24)
  def __init__(self, text):
    # 预设宽度 可以修改成你需要的图片宽度
    self.width = 600
    # 文本
    self.text = text
    # 段落 , 行数, 行高
    self.duanluo, self.note_height, self.line_height = self.split_text()
  def get_duanluo(self, text):
    txt = Image.new('RGBA', (100, 100), (255, 255, 255, 0))
    draw = ImageDraw.Draw(txt)
    # 所有文字的段落
    duanluo = ""
    # 宽度总和
    sum_width = 0
    # 几行
    line_count = 1
    # 行高
    line_height = 0
    for char in text:
      width, height = draw.textsize(char, ImgText.font)
      sum_width += width
      if sum_width > self.width: # 超过预设宽度就修改段落 以及当前行数
        line_count += 1
        sum_width = 0
        duanluo += '\n'
      duanluo += char
      line_height = max(height, line_height)
    if not duanluo.endswith('\n'):
      duanluo += '\n'
    return duanluo, line_height, line_count
  def split_text(self):
    # 按规定宽度分组
    max_line_height, total_lines = 0, 0
    allText = []
    for text in self.text.split('\n'):
      duanluo, line_height, line_count = self.get_duanluo(text)
      max_line_height = max(line_height, max_line_height)
      total_lines += line_count
      allText.append((duanluo, line_count))
    line_height = max_line_height
    total_height = total_lines * line_height
    return allText, total_height, line_height


  def draw_text(self):
    """
    绘图以及文字
    :return:
    """
    note_img = Image.open("001.png").convert("RGBA")
    draw = ImageDraw.Draw(note_img)
    # 左上角开始
    x, y = 25, 25
    for duanluo, line_count in self.duanluo:
      draw.text((x, y), duanluo, fill=(0, 0, 0), font=ImgText.font)
      y += self.line_height * line_count
    note_img.save(i_str+".png")
if __name__ == '__main__':
  for i in range(65,108):  #起始行数与结束行数-1
    n = ImgText(detail(i))
    i_str = str(i)
    n.draw_text()
    i=i+1