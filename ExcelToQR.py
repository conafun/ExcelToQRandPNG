#encoding:utf-8
from PIL import Image
import qrcode
import pygame
from openpyxl import load_workbook
pygame.init()  #初始化ygame
#读取excel单元格内容
wb=load_workbook('Sample.xlsx')
sheet=wb.worksheets[0]
sh=sheet.cell(1,1).value #(x,y)代表x行，y列
sh2=sheet.cell(1,2).value
sh3=sheet.cell(1,3).value
sh4=sheet.cell(1,4).value
sh5=sheet.cell(1,5).value
sh6=sheet.cell(1,6).value
#把excel内读取到的文字变成图片，生成二维码抬头
font=pygame.font.SysFont('FangSong',26)
ftext=font.render(sh+sh2,True,(0,0,0),(255,255,255))
pygame.image.save(ftext,'image.png')
#根据需要，收集不同单元格内信息，生产QR码
icon = Image.open("image.png")
img=qrcode.make(sh+sh2+sh3+sh4+sh5+sh6)
img2=img.paste(icon,(0, 0), mask=None)
with open('test.png', 'wb') as f:
    img.save(f)